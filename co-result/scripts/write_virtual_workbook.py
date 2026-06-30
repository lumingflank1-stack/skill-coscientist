#!/usr/bin/env python3
"""Write a simulated result-data workbook from a JSON spec.

This script is intentionally generic: Codex can generate a task-specific JSON
spec, then use this utility to apply workbook styling, four-decimal rounding,
and optional minimum-variance checks for simulated replicate groups.
"""

from __future__ import annotations

import argparse
import json
import math
import random
from pathlib import Path
from statistics import mean, stdev
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_SEED = 20260630


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True, help="JSON workbook spec")
    parser.add_argument("--output", required=True, help="Output .xlsx path")
    parser.add_argument("--decimals", type=int, default=None, help="Numeric decimal places; defaults to spec or 4")
    parser.add_argument("--seed", type=int, default=DEFAULT_SEED, help="Deterministic jitter seed")
    return parser.parse_args()


def as_rows(sheet: dict[str, Any]) -> tuple[list[str], list[list[Any]]]:
    columns = sheet.get("columns")
    rows = sheet.get("rows", [])
    if not columns:
        raise ValueError(f"Sheet {sheet.get('name')} is missing columns")

    normalized_rows: list[list[Any]] = []
    for row in rows:
        if isinstance(row, dict):
            normalized_rows.append([row.get(col) for col in columns])
        elif isinstance(row, list):
            if len(row) != len(columns):
                raise ValueError(f"Row length mismatch in sheet {sheet.get('name')}: {row}")
            normalized_rows.append(row)
        else:
            raise TypeError(f"Unsupported row type in sheet {sheet.get('name')}: {type(row)}")
    return list(columns), normalized_rows


def is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool) and math.isfinite(value)


def round_numbers(value: Any, decimals: int) -> Any:
    if is_number(value):
        return round(float(value), decimals)
    return value


def row_key(row: dict[str, Any], group_by: list[str]) -> tuple[Any, ...]:
    return tuple(row.get(col) for col in group_by)


def ensure_min_variance(spec: dict[str, Any], seed: int) -> None:
    rng = random.Random(seed)
    checks = spec.get("variance_checks", [])
    if not checks:
        return

    sheets_by_name = {sheet["name"]: sheet for sheet in spec.get("sheets", [])}
    for check in checks:
        sheet_name = check["sheet"]
        group_by = check.get("group_by", [])
        value_column = check["value_column"]
        min_sd = float(check.get("min_sd", 0.0))
        nonnegative = bool(check.get("nonnegative", True))
        sheet = sheets_by_name.get(sheet_name)
        if sheet is None:
            raise ValueError(f"Variance check references missing sheet: {sheet_name}")

        columns, rows = as_rows(sheet)
        if value_column not in columns:
            raise ValueError(f"Variance check value column missing in {sheet_name}: {value_column}")
        value_idx = columns.index(value_column)
        row_dicts = [dict(zip(columns, row)) for row in rows]

        groups: dict[tuple[Any, ...], list[int]] = {}
        for idx, row in enumerate(row_dicts):
            value = row.get(value_column)
            if is_number(value):
                groups.setdefault(row_key(row, group_by), []).append(idx)

        for indices in groups.values():
            if len(indices) < 3:
                continue
            values = [float(row_dicts[i][value_column]) for i in indices]
            current_sd = stdev(values) if len(set(values)) > 1 else 0.0
            if current_sd >= min_sd:
                continue

            center = mean(values)
            raw_offsets = [rng.gauss(0, 1) for _ in indices]
            offset_mean = mean(raw_offsets)
            centered = [x - offset_mean for x in raw_offsets]
            centered_sd = stdev(centered) if len(set(centered)) > 1 else 1.0
            scaled = [x / centered_sd * min_sd for x in centered]
            for row_i, offset in zip(indices, scaled):
                new_value = center + offset
                if nonnegative:
                    new_value = max(0.0, new_value)
                row_dicts[row_i][value_column] = new_value

        sheet["rows"] = [[row.get(col) for col in columns] for row in row_dicts]


def add_readme(wb: Workbook, spec: dict[str, Any]) -> None:
    ws = wb.active
    ws.title = "README"
    rows = [
        ("Item", "Value"),
        ("Workbook title", spec.get("workbook_title", "Virtual result-data workbook")),
        ("Source mode", "See the accompanying Markdown report and result_source_ledger.md for whether values are input, assumed, virtual, or requirements-only."),
        ("Allowed use", "Manuscript planning, figure prototyping, and real-data table templates."),
        ("Forbidden use", "Final statistical inference, submission, preprint, grant report, or any claim as measured data."),
        ("Generated by", "co-result/scripts/write_virtual_workbook.py"),
    ]
    for row in rows:
        ws.append(row)
    style_sheet(ws)
    ws["B3"].font = Font(bold=True)


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 8
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), 48))
        ws.column_dimensions[letter].width = max_len + 2


def write_sheet(wb: Workbook, sheet: dict[str, Any], decimals: int) -> None:
    name = sheet["name"]
    if name == "README":
        name = "README_extra"
    ws = wb.create_sheet(name[:31])
    columns, rows = as_rows(sheet)
    ws.append(columns)
    for row in rows:
        ws.append([round_numbers(value, decimals) for value in row])
    style_sheet(ws)


def main() -> None:
    args = parse_args()
    spec_path = Path(args.input)
    out_path = Path(args.output)
    spec = json.loads(spec_path.read_text(encoding="utf-8"))
    decimals = args.decimals if args.decimals is not None else int(spec.get("decimal_places", 4))

    ensure_min_variance(spec, args.seed)

    wb = Workbook()
    add_readme(wb, spec)
    for sheet in spec.get("sheets", []):
        write_sheet(wb, sheet, decimals)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(str(out_path))


if __name__ == "__main__":
    main()
